# modules
import bs4
import requests
from openpyxl import load_workbook, Workbook
import os
import re
flag = False
for i in os.listdir(os.getcwd()):
    if i.split('.')[-1] == 'xlsx':
        flag = True
print(os.getcwd())
if not flag:
    wb = Workbook()
    wb.save(os.path.join(os.getcwd(), 'table.xlsx'))


# Excel
wb = load_workbook('table.xlsx') # load table


sheet = wb['Sheet'] # load list

# Browser
url = 'https://wooordhunt.ru'
# const
dictWithWords = {}
number = 1
ASCII_code_word = 65
ASCII_code_translate = 66
ASCII_word_noun = 67
ASCII_word_verb = 68
sheet['A1'].value = 'Слово'
sheet['B1'].value = 'Перевод'

##############################################
while sheet['A' + str(number)].value != None:
    number += 1
    dictWithWords[sheet['A' + str(number)].value] = sheet['B' + str(number)].value
print('''Если вы хотите остановить выполнение программы введите слово: STOP''')
while True:
    word = input('Введите слово:\n')
    while True:
        if word == 'STOP':
            print('Программа остановлена')
            word = input('Для продолжения работы введите CONTINUE\n')
            if word == 'CONTINUE':
                word = input('Введите слово:\n')
        else:
            break
##############################################
    while True:
        if word in dictWithWords.keys():
            print('ДАННОЕ СЛОВО УЖЕ ЕСТЬ В СЛОВАРЕ')
            print('ПОВТОРИМ: {0} => {1}'.format(word, dictWithWords[word]))
            word = input('Введите слово:\n')
        else:
            break



    res = requests.get(url + '/word/' + word)
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    try:
        wordTranslate = ', '.join(soup.select('.t_inline_en')[0].getText().split(', ')[0:2])
    except:
        wordTranslate = soup.select('.t_inline_en')[0].getText().split(', ')[0]
    print('{0} => {1}'.format(word, wordTranslate))

    dictWithWords[word] = wordTranslate



    # Excel
    sheet[chr(ASCII_code_word) + str(number)].value = word # write value in table
    sheet[chr(ASCII_code_translate) + str(number)].value = wordTranslate

    wb.save('table.xlsx')
    number += 1

